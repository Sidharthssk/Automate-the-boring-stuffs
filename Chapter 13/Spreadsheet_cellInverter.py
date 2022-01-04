import openpyxl,sys

if len(sys.argv)==2:
    wb1 = openpyxl.load_workbook(sys.argv[1])
    sheet1 = wb1.active

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active

    for i in range(1,sheet1.max_row+1):
        for j in range(1,sheet1.max_column+1):
            
            sheet2.cell(row=j,column=i).value = sheet1.cell(row=i,column=j).value

    print('Data modified....')
    wb2.save('test.xlsx')