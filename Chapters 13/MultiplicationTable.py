
import openpyxl,sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter 

wb = openpyxl.Workbook()
sheet = wb.active
fontObj = Font(bold=True)

if len(sys.argv)>1:
    n = int(sys.argv[1])
    for i in range(1,n+1):
        sheet.cell(row=1,column=i+1).font = fontObj
        sheet.cell(row=i+1,column=1).font = fontObj
        sheet.cell(row=1,column=i+1).value=i
        sheet.cell(row=i+1,column=1).value=i
        sheet.cell(row=2,column=i+1).value=i
        sheet.cell(row=i+1,column=2).value=i
    
    for i in range(3,n+2):
        for j in range(3,n+2):
            n1 = int(sheet.cell(row=i,column=2).value)
            n2 = int(sheet.cell(row=1,column=j).value)
            product = n1*n2
            sheet.cell(row=i,column=j).value=product
        
    
wb.save('Multiplication.xlsx')