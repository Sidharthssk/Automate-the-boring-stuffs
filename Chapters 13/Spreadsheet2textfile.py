import openpyxl,os


wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active

for j in range(1,sheet.max_column+1):
    file = open(f'file{j}.txt','w')
    for i in range(1,sheet.max_row+1):
        file.write(str(sheet.cell(row=i,column=j).value))
        file.write('/n')
    file.close()