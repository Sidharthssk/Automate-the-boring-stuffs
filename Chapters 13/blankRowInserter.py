import openpyxl,sys

if len(sys.argv)>3:
    n = int(sys.argv[1])
    m = int(sys.argv[2])

    wb1 = openpyxl.load_workbook(sys.argv[3])
    sheet1 = wb1.active

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
  
    for i in range(1,sheet1.max_row):
        for j in range(1,sheet1.max_column):
            if i >= n:
                sheet2.cell(row=i+m,column=j).value = sheet1.cell(row=i,column=j).value
            else :
                sheet2.cell(row=i,column=j).value = sheet1.cell(row=i,column=j).value
    print('Saving to the file')
    wb2.save(sys.argv[3]) 
    
